import os
import pandas as pd
from datetime import datetime
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, PatternFill, Alignment


def scan_files(folder_path, min_size_mb=10, extensions=None):
    """
    File scanner (2 modes)
      extensions specified -> extract only those extensions
      extensions=None      -> extract all files (skip cache directories)
    """
    skip_dirs = {".gradle", ".conda", "node_modules", ".cache", "__pycache__",
                 ".venv", "venv", ".git", ".tox", ".mypy_cache"}
    # Normalize extensions (tuple, lowercase, dotted)
    if extensions is not None:
        if isinstance(extensions, str):
            extensions = (extensions,)
        extensions = tuple(e.lower() if e.startswith(".") else f".{e.lower()}" for e in extensions)

    data = []
    for root, dirs, files in os.walk(folder_path):
        dirs[:] = [d for d in dirs if d not in skip_dirs]
        for file in files:
            # Extension filter
            if extensions is not None and not file.lower().endswith(extensions):
                continue
            file_path = os.path.join(root, file)
            try:
                size_mb = os.path.getsize(file_path) / (1024 * 1024)
                if size_mb >= min_size_mb:
                    ext = os.path.splitext(file)[1].lower()
                    modified_time = datetime.fromtimestamp(
                        os.path.getmtime(file_path)
                    ).strftime("%Y-%m-%d %H:%M")
                    data.append([file, ext, root, modified_time, round(size_mb, 1)])
            except (FileNotFoundError, OSError):
                continue

    df = pd.DataFrame(data, columns=["File", "Ext", "Folder", "Date", "MB"])
    df.sort_values(by="MB", ascending=False, inplace=True)
    return df.reset_index(drop=True)


def to_excel(df, output_path="scan_result.xlsx"):
    """Export DataFrame to Excel with autofilter, auto-adjusted column widths, and Yu Gothic UI font."""
    # Reorder columns and add row number
    out = df[["MB", "Date", "Ext", "File", "Folder"]].copy()
    out["Date"] = pd.to_datetime(out["Date"]).dt.strftime("%Y/%m/%d")
    out.insert(0, "", [f"{i:02d}" for i in range(1, len(out) + 1)])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Files")
        ws = writer.sheets["Files"]

        # Autofilter
        ws.auto_filter.ref = ws.dimensions
        # Freeze top row
        ws.freeze_panes = "A2"

        # White fill for all empty cells (Normal style)
        theme_fill = PatternFill(patternType="solid", fgColor=Color(theme=0))
        writer.book._named_styles["Normal"].fill = theme_fill

        # White fill + font for data cells
        font = Font(name="Yu Gothic UI", size=11)
        white_fill = PatternFill(patternType="solid", fgColor=Color(theme=0))
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.fill = white_fill
                
        # Header row: left-aligned
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="left")

        # MB column number format "0.0"
        mb_col = list(out.columns).index("MB") + 1
        for row in ws.iter_rows(min_row=2, min_col=mb_col, max_col=mb_col):
            for cell in row:
                cell.number_format = "0.0"

        # Auto-adjust column widths
        for i, col in enumerate(out.columns, 1):
            max_len = max(out[col].astype(str).str.len().max(), len(col)) + 2
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max_len, 80)


if __name__ == "__main__":
    target = r"C:\Users\ky090"

    # --- Mode 1: Specific extensions only ---
    extensions = (".xls", ".xlsx", ".xlsm", ".xlsb", ".csv", ".pdf", ".ipynb")
    df = scan_files(target, min_size_mb=0.01, extensions=extensions)

    # --- Mode 2: All files ---
    # df = scan_files(target, min_size_mb=1)

    output_path = "Scan.xlsx"
    to_excel(df, output_path)
    print(f"Output: {os.path.abspath(output_path)}")
    display(df.head(7))
